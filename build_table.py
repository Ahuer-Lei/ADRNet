import os
import torch 
import json
import openpyxl
from openpyxl.styles import Font,Alignment
import torch.nn as nn
from datetime import datetime


address_dict={ "result_save_path":'./result.xlsx'

}


# 将实验结果输出为xlsx文件
def set_result(result,method_dict):
    '''
    将实验结果输出为xlsx文件
    method : 实验方法
    '''

    data_time=str(datetime.now())[0:19]
    experiment_result=[data_time,method_dict['method'],method_dict['id'],method_dict['batch_size'],method_dict['epoch']]
    
    if not os.path.exists(address_dict['result_save_path']):
        wb=openpyxl.Workbook()
        ws=wb.active
        ws.title='experiment_result'
        table_head=['日期','实验方法','实验编号','batch_size','epoch']
        for item in result.items():
            table_head.append(item[0])
            experiment_result.append(item[1])
        ws.append(table_head)
        ws.append(experiment_result)
    else:
        wb=openpyxl.load_workbook(address_dict['result_save_path'])
        ws=wb['experiment_result']
        for item in result.items():
            experiment_result.append(item[1])
        ws.append(experiment_result)
    
    wb.save(filename=address_dict['result_save_path']) 
    wb.close()
    set_font_style(address_dict['result_save_path'])

# def experiment_arrange(arrangement_dict):
#     '''
#     将实验结果输出为xlsx文件
#     method : 实验方法
#     '''
#     # method,file_path,serve,screen,gpu,epoch,lr,batch_size,warmup,total_epoch=arrangement
#     arrangement=arrangement_dict.values()
#     data_time=str(datetime.now())[0:19]
#     experiment_arrangement=[data_time]
#     for i in arrangement:
#         experiment_arrangement.append(i)
#     if not os.path.exists(address_dict['experiment_arrangement_path']):
#         wb=openpyxl.Workbook()
#         ws=wb.active
#         ws.title='experiment_arrangement'
#         table_head=['日期','实验方法','模型文件','实验编号','Over all Acc','Average Acc','服务器','screen环境名','gpu编号','加载权重','epoch','lr','momentum','batch_size','warmup','trained_epoch','total_epoch']
#         ws.append(table_head)
#         ws.append(experiment_arrangement)
#     else:
#         wb=openpyxl.load_workbook(address_dict['experiment_arrangement_path'])
#         ws=wb['experiment_arrangement']
#         ws.append(experiment_arrangement)
#     
#     wb.save(filename=address_dict['experiment_arrangement_path']) 
#     wb.close()
#     set_font_style(address_dict['experiment_arrangement_path'])

# def write_arrangement(message_list,row=None,save_note=None,id=None):
#     '''
#     message_list:[(column,info)]
#     '''
#     wb=openpyxl.load_workbook(address_dict['experiment_arrangement_path'])
#     ws=wb['experiment_arrangement']
#     if row is None:
#         method=ws['B']
#         
#         for item in method:
#             if item.value==save_note and ws[item.row][3].value==id:
#                 row=item.row
#                 break
# 
#     if len(message_list)>0 and row:
#         for items in message_list:
#             column=items[0]
#             info=items[1]
#             ws[row][column].value=info
#         wb.save(filename=address_dict['experiment_arrangement_path'])
#     wb.close()



def set_font_style(path):
    '''
    调整excel字体和格式
    '''
    if os.path.exists(path):
        wb=openpyxl.load_workbook(path)
       
        ws=wb['experiment_result']
    
        font = Font(name="微软雅黑",size=10)
        align = Alignment(horizontal="center",vertical="center")
        font_=Font(bold=True)
        max_rows = ws.max_row  # 获取最大行
        max_columns = ws.max_column  # 获取最大列

        # openpyxl的下标从1开始
        for i in range(1, max_rows + 1):
            for j in range(1, max_columns + 1):
                ws.cell(i, j).alignment = align
                ws.cell(i,j).font=font
        if path[-11:-5]=='result':
            for i in range(2, max_rows + 1):
                for j in range(5, max_columns + 1):
                    ws.cell(i,j).number_format='0.0000'
            
        else:
            for i in range(2, max_rows + 1):
                for j in [6,7,8,9]:
                    ws.cell(i,j).number_format='0.0000'
            
        for j in range(1, max_columns + 1):
            ws.cell(1,j).font=font_

        wb.save(path)
        wb.close()
# if __name__=='__main__':

#     result={'ir_self':0.76,'vi_self':0.89,'disp_error':0.89}
#     method_dict={'method':'dasfFASDFDSFASD','id':3,'batch_size':30,'epoch':300}
#     set_result(result=result,method_dict=method_dict)
#     print('hello')
    

    
